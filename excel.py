import requests
from bs4 import BeautifulSoup
import time
import json
import re
import os
import subprocess
from openpyxl import Workbook
from openpyxl import load_workbook
import pyautogui
import datetime
from pywinauto import Application
import pyperclip

# Move the mouse to the specified (x, y) coordinates
def move_and_click(x, y):
    pyautogui.moveTo(x, y)  
    time.sleep(0.5)
    pyautogui.click()

base_url = 'https://www.kohls.com'
headers = {
    'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36'
}

#Prompt
search = input("Seach Kohls For:")
search_input = search.replace(" ","+")
page_url = "https://www.kohls.com/search.jsp?submit-search=&search="+search_input
print(page_url)
workbook = Workbook()
sheet = workbook.active

page_number = 1

more_pages = True
product_number_total=0

cells=0
items=0
information=[None, None]
while more_pages:
    try:
        r = requests.get(page_url, headers=headers)
        r.raise_for_status()  # raises an exception if the request was not successful
        # print(r.content)
    except requests.exceptions.RequestException as e:
        print(e)
        print('An error occurred while making the request')
    else:
        # request was successful
        soup = BeautifulSoup(r.content, 'lxml')

    # Makes product list equal to all divs with class products_grid
    product_list = soup.find_all('li', class_='products_grid')
    for product in product_list:
        
        product_number = product_list.index(product)+1+ product_number_total
        print("product #",product_number,": Item #", items,": Cell #", cells)
        # Extract the name of the product
        name = product.find('div', class_='prod_nameBlock').text
        name = name.strip()
        #print(name)
        #print()

        # Extract the link to the product page
        link = product.find('div', class_='prod_img_block').find('a')['href']  
        link=base_url+link
        #print(link)

        try:
            UPCresponse = requests.get(link, headers=headers)
            UPCresponse.raise_for_status()
        except requests.exceptions.RequestException as e:
            print(e)
            print('An error occurred while making the request')
        else:
            # request was successful
            # find the productV2JsonData variable in the JavaScript code
            pattern = r'var productV2JsonData = ({.*?});'
            match = re.search(pattern, UPCresponse.text, re.DOTALL)

            if match:
                upc_price=[]
                # extract the JSON data from the matched group
                json_data = match.group(1)

                # parse the JSON data
                product_data = json.loads(json_data)

                # do something with the product data
                #upcs=0
                
                for product in product_data['SKUS']:
                    #upcs+=1
                    upc = product['UPC']['ID']
                    price = product['price']['lowestApplicablePrice']#['minPrice']
                    availability = product['availability']
                    upc = str(upc)
                    price = str(price)
                    #print(price)
                    if availability != "Out of Stock":
                        #information.append({'upc':upc, 'price':price})
                        #information.append(upc)
                        #information.append(price)
                        #information.append(upc_price)
                        #user = upc, price
                        sheet.append([upc, price])
                        cells +=2
                        items +=1
            else:
                print('Product data not found in JavaScript code')

    # Check if there is a next page
    pagination_block = soup.find('div', class_='prodsPagination_block fr')
    if pagination_block:
        # Find the link to the next page
        print('There is a next page')
        page_number += 1
        link_element = soup.find('link', rel='next')
        if link_element == None:
            #worksheet.insert_row(information, 2)
            more_pages = False
            break
        else:
            next_page_link = link_element['href']
            print(next_page_link)
            #cells+=1
            #worksheet.insert_row(next_page_link, 2)
            if next_page_link:
                # Update the page number and URL
                page_number += 1
                product_number_total = product_number
                page_url = base_url + next_page_link
                # Make a request to the next page
                r = requests.get(page_url, headers=headers)
                soup = BeautifulSoup(r.content, 'lxml')
            else:
                # There is no next page
                more_pages = False
    else:
        # There is no pagination block, so there must be only one page
        more_pages = False

print("product #",product_number,": Item #", items,": Cell #", cells)

# Save the workbook
workbook.save(filename=search + ".xlsx")

# Load the workbook to refresh the Excel application
loaded_workbook = load_workbook(filename=search+ ".xlsx")
loaded_workbook.close()

# Open the workbook with the default application
###os.startfile(search+ ".xlsx")

os.startfile(search + ".xlsx")

# Wait for Excel to open
time.sleep(2)    

app = Application().connect(title_re=".*Excel", found_index=0)
# Maximize the Excel window
window = app.window(title_re=".*Excel", found_index=0)
window.Maximize()

move_and_click(1850, 50)
move_and_click(1850, 100)
time.sleep(3)

# Get the contents of the clipboard
sheet_link = pyperclip.paste()
print(sheet_link)

pyautogui.hotkey("alt", "f4")
pyautogui.hotkey("alt", "f4")

# Creates a subject for the email
# Get the current date
current_date = datetime.datetime.now().strftime("%m-%d")
subject = f'{search}, {current_date}'

# Open the file location in the file explorer
# Specify the file path of the Excel file
excel_file_path = os.path.abspath(search + ".xlsx")
# Construct the command to open the mail app with the attachment
excel_file = "file://"+excel_file_path
command = 'start "" "mailto:ccuellar418@gmail.com?subject={}&body={}"'.format(subject, sheet_link)

# Open the mail app with the attachment
subprocess.Popen(command, shell=True)
time.sleep(5)
pyautogui.hotkey("ctrl", "enter")

print("POGGERS!")